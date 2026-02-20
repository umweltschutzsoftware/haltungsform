"""Parser für die profarm XLSX-Datei (Blatt 'Daten')."""

import io

import openpyxl

from src.models import IstZustandRow, PlanZustandRow


def parse_xlsx(file_bytes: bytes) -> dict:
    """Parse die profarm XLSX-Datei und extrahiere Betriebsdaten.

    Erwartetes Format (Blatt 'Daten'):
    - B1: Straße, B2: Hausnummer, B3: PLZ, B4: Ort, B5: Projektnummer, B6: E-Mail
    - Zeilen 9-50: Betriebseinheiten
      - A: BE-Nr, B: Tierart (Ist), C: Tierplätze (Ist), D: Ausführung (Ist),
        E: Kamine, F: S.d.T., G: Tierart (Ziel), H: Tierplätze (Ziel), I: Ausführung (Ziel)

    Returns:
        dict mit Schlüsseln: strasse, hausnummer, plz, ort, projektnummer,
        email, ist_zustand, plan_zustand
    """
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)

    # Blatt "Daten" suchen
    if "Daten" in wb.sheetnames:
        ws = wb["Daten"]
    else:
        # Fallback: erstes Blatt verwenden
        ws = wb.worksheets[0]

    # Adressdaten auslesen
    result = {
        "strasse": _cell_str(ws, "B1"),
        "hausnummer": _cell_str(ws, "B2"),
        "plz": _cell_str(ws, "B3"),
        "ort": _cell_str(ws, "B4"),
        "projektnummer": _cell_str(ws, "B5"),
        "email": _cell_str(ws, "B6"),
    }

    # Betriebseinheiten auslesen (Zeilen 9-50)
    ist_zustand = []
    plan_zustand = []

    for row_num in range(9, 51):
        be_nr = ws.cell(row=row_num, column=1).value  # A
        tierart_ist = ws.cell(row=row_num, column=2).value  # B

        # Zeile überspringen wenn BE-Nr und Tierart leer
        if be_nr is None and tierart_ist is None:
            continue

        be_nr_str = str(be_nr) if be_nr is not None else ""
        tierplaetze_ist = ws.cell(row=row_num, column=3).value  # C
        ausfuehrung_ist = ws.cell(row=row_num, column=4).value  # D
        kamine = ws.cell(row=row_num, column=5).value  # E
        sdt = ws.cell(row=row_num, column=6).value  # F

        tierart_ziel = ws.cell(row=row_num, column=7).value  # G
        tierplaetze_ziel = ws.cell(row=row_num, column=8).value  # H
        ausfuehrung_ziel = ws.cell(row=row_num, column=9).value  # I

        ist_zustand.append(IstZustandRow(
            be_nr=be_nr_str,
            tierart=str(tierart_ist) if tierart_ist else "Mastschweine",
            tierplaetze=int(tierplaetze_ist) if tierplaetze_ist else 0,
            ausfuehrung=_normalize_ausfuehrung(ausfuehrung_ist),
            kamine=str(kamine) if kamine else "Nein",
            stand_der_technik=str(sdt) if sdt else "Nein",
        ))

        plan_zustand.append(PlanZustandRow(
            be_nr=be_nr_str,
            tierart=str(tierart_ziel) if tierart_ziel else "Mastschweine",
            tierplaetze=int(tierplaetze_ziel) if tierplaetze_ziel else 0,
            ausfuehrung=_normalize_ausfuehrung(ausfuehrung_ziel),
        ))

    result["ist_zustand"] = ist_zustand
    result["plan_zustand"] = plan_zustand

    return result


def _cell_str(ws, cell_ref: str) -> str:
    """Liest einen Zellwert als String."""
    val = ws[cell_ref].value
    if val is None:
        return ""
    return str(val).strip()


def _normalize_ausfuehrung(val) -> str:
    """Normalisiert Ausführungs-Werte aus der XLSX.

    Die XLSX enthält z.B. '1 - Zwangsbelüfteter Stall'.
    Wir behalten das vollständige Format bei.
    """
    if val is None:
        return "1 - Zwangsbelüfteter Stall"

    val_str = str(val).strip()

    # Falls nur eine Zahl angegeben ist
    ausfuehrung_map = {
        "1": "1 - Zwangsbelüfteter Stall",
        "2": "2 - Zwangsbelüfteter Stall mit Auslauf",
        "3": "3 - Außenklimastall",
        "4": "4 - Außenklimastall mit Auslauf",
    }
    if val_str in ausfuehrung_map:
        return ausfuehrung_map[val_str]

    return val_str
